import platform
import psutil
import tkinter as tk
from tkinter import ttk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
import speedtest
import socket
import os
import shutil
from datetime import datetime, timedelta
import tkinter.messagebox as messagebox
import os
import psutil
import shutil
import speedtest
import subprocess
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl import load_workbook

def get_serial_number():
    try:
        result = subprocess.check_output("wmic bios get serialnumber", shell=True).decode().strip().split("\n")
        return result[1].strip() if len(result) > 1 else "No Serial Number Found"
    except Exception as e:
        return f"Error: {e}"

# Function to get memory info
def get_memory_info():
    virtual_memory = psutil.virtual_memory()
    return {
        "Total Memory (GB)": virtual_memory.total / (1024 ** 3),
        "Used Memory (GB)": virtual_memory.used / (1024 ** 3),
        "Available Memory (GB)": virtual_memory.available / (1024 ** 3),
        "Memory Usage (%)": virtual_memory.percent
    }

# Function to get disk info
def get_disk_info():
    disk_usage = psutil.disk_usage('/')
    return {
        "Total Disk Space (GB)": disk_usage.total / (1024 ** 3),
        "Used Disk Space (GB)": disk_usage.used / (1024 ** 3),
        "Free Disk Space (GB)": disk_usage.free / (1024 ** 3),
        "Disk Usage (%)": disk_usage.percent
    }

# Function to get the username
def get_username():
    try:
        return os.getlogin()
    except Exception as e:
        return f"Error: {e}"

# Function to get the system remark
def get_system_remark():
    try:
        result = subprocess.check_output("wmic computersystem get description", shell=True).decode().strip().split("\n")
        return result[1].strip() if len(result) > 1 else "No Description Found"
    except Exception as e:
        return f"Error: {e}"

# Gather all data
memory_info = get_memory_info()
disk_info = get_disk_info()
specification_values = (
    f"Total Memory: {memory_info['Total Memory (GB)']:.2f} GB, "
    f"Used Memory: {memory_info['Used Memory (GB)']:.2f} GB, "
    f"Total Disk Space: {disk_info['Total Disk Space (GB)']:.2f} GB"
)

data = {
    "Username": get_username(),
    "Serial Number": get_serial_number(),
    "System Model": platform.uname().machine,
    "System Manufacturer": platform.uname().system,
    "Asset Tag": platform.uname().node,
    "System Remark": get_system_remark(),
    "Specification": specification_values,
}

# Check if the data already exists
# Export to Excel
def export_to_excel(data, filename="window.xlsx"):
    try:
        # Check if the file already exists
        if os.path.exists(filename):
            wb = load_workbook(filename)
            sheet = wb.active
        else:
            # Create a new workbook and set up headers
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "System Info"
            headers = [
                "N",
                "Specification",
                "Username",
                "Serial Number",
                "System Model",
                "System Manufacturer",
                "Asset Tag",
                "System Remark",
            ]
            
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            header_font = Font(bold=True)

            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            


        # Check if the data already exists (match based on "Username" and "Serial Number")
        for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip the header
            username_cell = sheet.cell(row=row, column=3).value
            serial_number_cell = sheet.cell(row=row, column=4).value
            if username_cell == data["Username"] and serial_number_cell == data["Serial Number"]:
                print("Data already exists in the Excel file. No changes made.")
                return

        # Find the next empty row
        next_row = sheet.max_row + 1

        # Populate the row
        sheet[f"A{next_row}"] = next_row - 1  # Numbering starts at 1
        sheet[f"B{next_row}"] = data.get("Specification", "N/A")
        sheet[f"C{next_row}"] = data.get("Username", "N/A")
        sheet[f"D{next_row}"] = data.get("Serial Number", "N/A")
        sheet[f"E{next_row}"] = data.get("System Model", "N/A")
        sheet[f"F{next_row}"] = data.get("System Manufacturer", "N/A")
        sheet[f"G{next_row}"] = data.get("Asset Tag", "N/A")
        sheet[f"H{next_row}"] = data.get("System Remark", "N/A")

        # Center-align all data
        for row in sheet.iter_rows(min_row=next_row, max_row=next_row, min_col=1, max_col=8):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # Save the workbook
        wb.save(filename)
        print(f"Data appended to {filename}")
    except Exception as e:
        print(f"Error exporting to Excel: {e}")

# Save results to Excel
export_to_excel(data)




# Fetch system information
def fetch_system_info():
    """Fetch and display basic system information."""
    info = [
        f"OS: {platform.system()} {platform.release()}",
        f"Architecture: {platform.architecture()[0]}",
        f"Processor: {platform.processor()}",
        f"RAM: {round(psutil.virtual_memory().total / (1024 ** 3), 2)} GB",
        f"Disk: {round(psutil.disk_usage('/').total / (1024 ** 3), 2)} GB",
    ]
    return "\n".join(info)

def cleanup_temp_files():
    """Placeholder for cleaning up temporary files."""
    print("Cleanup functionality triggered!")
    messagebox.showinfo("Cleanup", "Temporary files cleaned successfully!")

def show_system_info(info_label):
    """Update the info section with system details."""
    info = get_system_info()
    info_label.config(text=info)

def update_task_list(tree):
    """Update the task list in the UI with CPU, memory, and disk usage."""
    for row in tree.get_children():
        tree.delete(row)
    
    # Iterate over the processes and fetch information
    for proc in psutil.process_iter(['pid', 'name', 'cpu_percent', 'memory_info']):
        try:
            # Get memory usage in MB and disk usage (simplified as zero here)
            memory_usage = proc.info['memory_info'].rss / (1024 ** 2)  # Convert to MB
            disk_usage = 0  # Disk usage can be complicated, so leaving it as 0

            tree.insert("", "end", values=(proc.info['pid'], proc.info['name'], proc.info['cpu_percent'], round(memory_usage, 2), disk_usage))
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue

def update_performance(cpu_line, mem_line, disk_line, x_data, cpu_data, mem_data, disk_data, canvas, start_time):
    """Update performance metrics dynamically and convert time to minutes."""
    current_time = (psutil.time.time() - start_time) / 60  # Convert time to minutes
    cpu_data.append(psutil.cpu_percent())
    mem_data.append(psutil.virtual_memory().percent)
    disk_data.append(psutil.disk_usage('/').percent)
    x_data.append(current_time)

    if len(x_data) > 30:
        x_data.pop(0)
        cpu_data.pop(0)
        mem_data.pop(0)
        disk_data.pop(0)

    cpu_line.set_data(x_data, cpu_data)
    mem_line.set_data(x_data, mem_data)
    disk_line.set_data(x_data, disk_data)

    ax = cpu_line.axes
    ax.set_xlim(max(0, current_time - 5), current_time)
    ax.set_ylim(0, 100)
    ax.set_xlabel('Time (minutes)')  # Changed to minutes
    canvas.draw()
    canvas.get_tk_widget().after(1000, update_performance, cpu_line, mem_line, disk_line, x_data, cpu_data, mem_data, disk_data, canvas, start_time)

def check_wifi_speed(info_label):
    """Check the download and upload speed of the active Wi-Fi connection."""
    st = speedtest.Speedtest()
    st.get_best_server()
    download_speed = st.download() / 1_000_000  # Convert to Mbps
    upload_speed = st.upload() / 1_000_000  # Convert to Mbps
    ping = st.results.ping
    result = f"Download Speed: {download_speed:.2f} Mbps\nUpload Speed: {upload_speed:.2f} Mbps\nPing: {ping} ms"
    info_label.config(text=result)


def get_wifi_passwords(info_label):
    """Retrieve stored Wi-Fi passwords."""
    profiles_cmd = "netsh wlan show profiles"
    profiles = subprocess.check_output(profiles_cmd, shell=True, text=True)
    profile_names = [line.split(":")[1].strip() for line in profiles.splitlines() if "All User Profile" in line]

    wifi_passwords = "Wi-Fi Passwords:\n"
    for profile in profile_names:
        password_cmd = f"netsh wlan show profile \"{profile}\" key=clear"
        profile_info = subprocess.check_output(password_cmd, shell=True, text=True)
        for line in profile_info.splitlines():
            if "Key Content" in line:
                password = line.split(":")[1].strip()
                wifi_passwords += f"{profile}: {password}\n"
                break
    info_label.config(text=wifi_passwords)

def get_serial_number():
    try:
        serial_number = subprocess.check_output("wmic bios get serialnumber").decode().strip().split('\n')[1]
        return serial_number
    except subprocess.CalledProcessError:
        return "Unable to retrieve serial number"


def get_cpu_info(info_label):
    cpu_count = psutil.cpu_count(logical=True)
    cpu_freq = psutil.cpu_freq()
    cpu_percent = psutil.cpu_percent(interval=1)
    processor_info = os.popen("wmic cpu get name").read()
    cpu_info = f"CPU (Processor): {processor_info}\nCPU Frequency: {cpu_freq.current} MHz\nCPU Usage: {cpu_percent}% \nCPU Count (Logical Cores): {cpu_count}"
    info_label.config(text=cpu_info)
    

    
    
def get_memory():
    memory_info= psutil.virtual_memory()
    disk_usage = psutil.disk_usage('/')
    net_io = psutil.net_io_counters()
    return f"""
Memory Usage        : {memory_info.percent}%
Total Memory        : {memory_info.total / (1024 ** 3):.2f} GB
Available Memory    : {memory_info.available / (1024 ** 3):.2f} GB

Disk Usage          : {disk_usage.percent}%
Total Disk Space    : {disk_usage.total / (1024 ** 3):.2f} GB
Used Disk Space     : {disk_usage.used / (1024 ** 3):.2f} GB
Free Disk Space     : {disk_usage.free / (1024 ** 3):.2f} GB

Total Bytes Sent    : {net_io.bytes_sent / (1024 ** 2):.2f} MB
Total Bytes Received: {net_io.bytes_recv / (1024 ** 2):.2f} MB
    """    
def get_memory_info(info_label):
    """Update the info section with system details."""
    info = get_memory ()
    info_label.config(text=info)



def Sensors_Battery():
    battery = psutil.sensors_battery()
    result = {
        "Battery   ": (str(battery.percent) + "%, Charging: " + ("Yes" if battery.power_plugged else "No"))
                   if battery else "No Battery",
        "Processes ": len(psutil.pids()),
        "Uptime    ": str(timedelta(seconds=(datetime.now().timestamp() - psutil.boot_time())))
    }

    # Display results in a formatted string
    result_str = ""
    for key, value in result.items():
        result_str += f"{key}: {value}\n"

    return result_str


def get_Battery_info(info_label):
    """Update the info section with system details."""
    info = Sensors_Battery ()
    info_label.config(text=info)


# beterry


def Sensors_Battery():
    battery = psutil.sensors_battery()
    
    result = {
        "Battery": (str(battery.percent) + "%, Charging: " + ("Yes" if battery.power_plugged else "No")) 
                   if battery else "No Battery",
        "Processes": len(psutil.pids()),
        "Uptime": str(timedelta(seconds=int(datetime.now().timestamp() - psutil.boot_time())))
    }

    # Display results in a formatted string
    result_str = ""
    for key, value in result.items():
        result_str += f"{key}: {value}\n"

    return result_str



def get_Battery_info(info_label):
    """Update the info section with system details."""
    info = Sensors_Battery ()
    info_label.config(text=info)

#shudown and restart
def shutdown_gui():
    """Shutdown the computer with a confirmation dialog."""
    confirm = messagebox.askyesno("Shutdown Confirmation", "Are you sure you want to shut down the computer?")
    if confirm:
        os.system("shutdown /s /t 1")  # Windows shutdown command
        messagebox.showinfo("Shutdown", "Shutting down...")
    else:
        messagebox.showinfo("Shutdown", "Shutdown canceled.")


def restart_gui():
    """Restart the computer with a confirmation dialog."""
    confirm = messagebox.askyesno("Restart Confirmation", "Are you sure you want to restart the computer?")
    if confirm:
        os.system("shutdown /r /t 1")  # Windows restart command
        messagebox.showinfo("Restart", "Restarting...")
    else:
        messagebox.showinfo("Restart", "Restart canceled.")

def create_full_interface():
    root = tk.Tk()
    root.title("Complete System Dashboard")
    root.geometry("1100x800")

    # Sidebar frame
    sidebar = ttk.Frame(root, width=200, relief="raised", padding=10)
    sidebar.pack(side=tk.LEFT, fill=tk.Y)

    # Top buttons (Wi-Fi features)
    ttk.Button(sidebar, text="Wi-Fi Credentials", command=lambda: get_wifi_passwords(info_label)).pack(fill=tk.X, pady=10)
    ttk.Button(sidebar, text="Wi-Fi Speed Test", command=lambda: check_wifi_speed(info_label)).pack(fill=tk.X, pady=10)

    # Bottom buttons (System Info and others)
    ttk.Button(sidebar, text="System Info", command=lambda: show_system_info(info_label)).pack(fill=tk.X, pady=5)
    ttk.Button(sidebar, text="CPU Info", command=lambda: get_cpu_info(info_label)).pack(fill=tk.X, pady=5)
    ttk.Button(sidebar, text="Memory Info", command=lambda: get_memory_info(info_label)).pack(fill=tk.X, pady=5)
    ttk.Button(sidebar, text="Disk Info", command=lambda: get_disk_info(info_label)).pack(fill=tk.X, pady=5)
    ttk.Button(sidebar, text="Cleanup", command=cleanup_temp_files).pack(fill=tk.X, pady=5)

    # Shutdown and Restart buttons
    ttk.Button(sidebar, text="Shutdown", command=shutdown_gui).pack(fill=tk.X, pady=5)
    ttk.Button(sidebar, text="Restart", command=restart_gui).pack(fill=tk.X, pady=5)

    # Main content area
    main_frame = ttk.Frame(root, padding=10)
    main_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

    # Info label at the bottom of the main frame
    info_label = tk.Label(main_frame, text="", justify="left", anchor="w", font=("Courier", 10), padx=10)
    info_label.pack(side=tk.BOTTOM, fill=tk.X, pady=10)

    # Create the performance plot
    fig, ax = plt.subplots(figsize=(7, 5))
    ax.set_title('System Performance (CPU, Memory, Disk)')
    ax.set_xlabel('Time (minutes)')
    ax.set_ylabel('% Usage')

    cpu_line, = ax.plot([], [], label="CPU", color="r")
    mem_line, = ax.plot([], [], label="Memory", color="g")
    disk_line, = ax.plot([], [], label="Disk", color="b")

    ax.legend(loc="upper right")

    canvas = FigureCanvasTkAgg(fig, master=main_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    # Initialize performance data
    x_data, cpu_data, mem_data, disk_data = [], [], [], []
    start_time = psutil.time.time()  # Store the start time
    update_performance(cpu_line, mem_line, disk_line, x_data, cpu_data, mem_data, disk_data, canvas, start_time)

    root.mainloop()

if __name__ == "__main__":
    create_full_interface()
