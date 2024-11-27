import os
import platform
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import load_workbook
import psutil

# Function to get the serial number
def get_serial_number():
    try:
        result = subprocess.check_output("wmic bios get serialnumber", shell=True).decode().strip().split("\n")
        return result[1].strip() if len(result) > 1 else "No Serial Number Found"
    except Exception as e:
        return f"Error: {e}"

# Function to get memory info
def get_memory_info():
    virtual_memory = psutil.virtual_memory()
    return {
        "Total Memory (GB)": virtual_memory.total / (1024 ** 3),
        "Used Memory (GB)": virtual_memory.used / (1024 ** 3),
        "Available Memory (GB)": virtual_memory.available / (1024 ** 3),
        "Memory Usage (%)": virtual_memory.percent
    }

# Function to get disk info
def get_disk_info():
    disk_usage = psutil.disk_usage('/')
    return {
        "Total Disk Space (GB)": disk_usage.total / (1024 ** 3),
        "Used Disk Space (GB)": disk_usage.used / (1024 ** 3),
        "Free Disk Space (GB)": disk_usage.free / (1024 ** 3),
        "Disk Usage (%)": disk_usage.percent
    }

# Function to get the username
def get_username():
    try:
        return os.getlogin()
    except Exception as e:
        return f"Error: {e}"

# Function to get the system remark
def get_system_remark():
    try:
        result = subprocess.check_output("wmic computersystem get description", shell=True).decode().strip().split("\n")
        return result[1].strip() if len(result) > 1 else "No Description Found"
    except Exception as e:
        return f"Error: {e}"

# Gather all data
memory_info = get_memory_info()
disk_info = get_disk_info()
specification_values = (
    f"Total Memory: {memory_info['Total Memory (GB)']:.2f} GB, "
    f"Used Memory: {memory_info['Used Memory (GB)']:.2f} GB, "
    f"Total Disk Space: {disk_info['Total Disk Space (GB)']:.2f} GB"
)

data = {
    "Username": get_username(),
    "Serial Number": get_serial_number(),
    "System Model": platform.uname().machine,
    "Operating system": platform.uname().system,
    "Asset Tag": platform.uname().node,
    "System Remark": get_system_remark(),
    "Specification": specification_values,
}

# Export to Excel
def export_to_excel(data, filename="window.xlsx"):
    try:
        # Check if the file already exists
        if os.path.exists(filename):
            wb = load_workbook(filename)
            sheet = wb.active
        else:
            # Create a new workbook and set up headers
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "System Info"
            headers = [
                "N",
                "Specification",
                "Username",
                "Serial Number",
                "System Model",
                "System Manufacturer",
                "Asset Tag",
                "System Remark",
            ]
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            header_font = Font(bold=True)

            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

        # Check if the data already exists (match based on "Username" and "Serial Number")
        for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip the header
            username_cell = sheet.cell(row=row, column=3).value
            serial_number_cell = sheet.cell(row=row, column=4).value
            if username_cell == data["Username"] and serial_number_cell == data["Serial Number"]:
                print("Data already exists in the Excel file. No changes made.")
                return

        # Find the next empty row
        next_row = sheet.max_row + 1

        # Populate the row
        sheet[f"A{next_row}"] = next_row - 1  # Numbering starts at 1
        sheet[f"B{next_row}"] = data.get("Specification", "N/A")
        sheet[f"C{next_row}"] = data.get("Username", "N/A")
        sheet[f"D{next_row}"] = data.get("Serial Number", "N/A")
        sheet[f"E{next_row}"] = data.get("System Model", "N/A")
        sheet[f"F{next_row}"] = data.get("Operating system", "N/A")
        sheet[f"G{next_row}"] = data.get("Asset Tag", "N/A")
        sheet[f"H{next_row}"] = data.get("System Remark", "N/A")

        # Center-align all data
        for row in sheet.iter_rows(min_row=next_row, max_row=next_row, min_col=1, max_col=8):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # Save the workbook
        wb.save(filename)
        print(f"Data appended to {filename}")
    except Exception as e:
        print(f"Error exporting to Excel: {e}")

# Save results to Excel
export_to_excel(data)

